#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''Funtions to manipulate exported Infa XML files to find nodes that match
with the desired xpath pattern specified in the components.yaml configuration
file and save the results into differents formats
such as: text (stdout), csv and Excel
'''

# standard libraries
import os
from string import Template
import csv
# third party libraries
from lxml import etree
import yaml
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# common libraries
from infa.common import get_exec_path, basename

# Long column width and long row height
LONG_COLUMN=90
LONG_ROW=60

class XmlInfaFormat:
	'''Functions to flat the structure of a XML node from infa file'''
	@staticmethod
	def __print_single_node(node, attr=None):
		'''Get the requested attribute(s) from the given node'''
		if node is None:
			return ''
		if attr is None:
			attr='$NAME'
		s=Template(attr)
		return s.safe_substitute(node.attrib)
	@staticmethod
	def expand_node(node, format=None):
		'''Traverse the structure to get the names (or the requested 
		attributes) from a node's hierarchy'''
		result=[]
		curnode=node
		format=format if format is not None else []
		while curnode is not None:
			tag=curnode.tag
			attr=None
			for item in format:
				attr=item[tag] if tag in item else None
			result.insert(0, XmlInfaFormat.__print_single_node(curnode,attr) )
			parent=curnode.getparent()
			if parent is not None and parent.tag == 'REPOSITORY':
				break
			curnode=parent
		return result
	@staticmethod
	def to_string(node, format=None):
		'''Given a node, get a string with the names of its predecessors'''
		result=XmlInfaFormat.expand_node(node, format)
		return '.'.join(result)
	@staticmethod
	def flat_it(node, format=None):
		'''Traverse the structure to get the name (or requested attributes) 
		from a node's hierarchy'''
		result=[]
		curnode=node
		while curnode is not None:
			attr=None
			if format is not None:
				tag=curnode.tag
				attr=format[tag] if tag in format else None
			result.insert(0, XmlInfaFormat.__print_single_node(curnode,attr) )
			parent=curnode.getparent()
			if parent is not None and parent.tag == 'REPOSITORY':
				break
			curnode=parent
		return result
	@staticmethod
	def headers(node):
		'''Given a node, get a list with the tags of its predecessors'''
		result=[]
		curnode=node
		while curnode is not None:
			result.insert(0, curnode.tag )
			parent=curnode.getparent()
			if parent is not None and parent.tag == 'REPOSITORY':
				break
			curnode=parent
		return result

def printit(root, xpath, format=None):
	'''Given a node (root), find the components that matches the requested 
	xpath definition using the specified format'''
	first=True
	result=[]
	for node in root.xpath(xpath):
		if first:
			result.append(XmlInfaFormat.headers(node))
			#print(XmlInfaFormat.headers(node))
			first=False
		result.append(XmlInfaFormat.flat_it(node, format))
		#print(XmlInfaFormat.flat_it(node, format))
	return result

def find_nodes_as_list(root, comp, parent=None, conf=None):
	'''Given a node, find the components that matches the requested definition
	within the provided configuration file. Optionally a parent component 
	could be specified'''
	data={}
	result=[]
	# if not set, use the default component configuration file
	conf=os.path.join(get_exec_path(),'config','components.yaml') if conf is None else conf
	# load component configuration
	with open(conf) as f:
		data = yaml.load(f, Loader=yaml.FullLoader)
	if parent is None:
		if comp not in data['components']:
			raise Exception(f"Component {comp} does not exists, please verify")
		xpath=data['components'][comp]
	else:
		if parent not in data['nested'] or comp not in data['nested'][parent]:
			raise Exception(f"Component {comp} does not exists, please verify")
		xpath=data['nested'][parent][comp]
	if type(xpath) is dict:
		format=xpath['format']
		xpath=xpath['xpath']
	else:
		format=None
	return printit(root, xpath, format)

def find_from_file(filename, comp, parent=None, conf=None, format='text', target=None):
	'''Find all the nodes of the given component (comp) into the root node or 
	as child of the specified parent (optional) and save it into the specified 
	format (text, csv or Excel) with the given name (target)
	'''
	# check if the specified file exists
	if not os.path.isfile(filename):
		raise Exception(f"File {filename} does not exists, please verify")
	# parse file
	print(f'Reading from {filename} ...')
	print(f'Parsing {filename} ...')
	tree=etree.parse(filename)
	root=tree.getroot()
	# check if is a valid exported Infa XML file
	if root.tag != 'POWERMART':
		raise Exception(f"The given file {filename} is not a valid exported INFA XML file, please verify")
	# find all the nodes of the given [parent.]comp
	print(f'Find {comp} into {filename} ...')
	try:
		result=find_nodes_as_list(root=root, comp=comp, parent=parent, conf=conf)
	except:
		print(f'Notice: Component [{comp}] not found, no files generated. Please verify')
		return
	if result is None or len(result)==0:
		print(f'Notice: No components {comp} found on {filename}, no files generated. Please verify')
		return
	if target is None:
		ext='.xlsx' if format=='excel' else '.csv'
		target=os.path.splitext(basename(filename))[0]+ext
	# TODO: enhance generating files in multiple formats
	format = 'text' if format not in ('text', 'csv', 'excel') else format
	print(f'Exporting to {format} ...')
	if format=='text':
		export_txt(result)
	elif format=='csv':
		export_csv(result, target)
	elif format=='excel':
		export_excel(result, target)
	# Done
	print('Done')

def as_text(value):
	'''retur the specified value as a string'''
	if value is None:
		return ""
	return str(value)

def export_excel(data, filename, sheetname=None):
	'''Given a dataset (as a list of lists), save it to the specified file
	with the specified sheetname as an Excel file'''
	has_long_cols=False
	if len(data)==0:
		return
	workbook = Workbook()
	sheet = workbook.active
	if sheetname is not None:
		sheet.title=sheetname
	for row in data:
		sheet.append(row)
	# freeze first column
	sheet.freeze_panes="A2"
	# add filter
	sheet.auto_filter.ref=sheet.dimensions
	# autosize (or kind of) columns
	for column_cells in sheet.columns:
		length = max(len(as_text(cell.value)) for cell in column_cells)+5
		if length > LONG_COLUMN:
			has_long_cols=True
			length = LONG_COLUMN
		sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
	# if long columns present, set column height
	if has_long_cols:
		for x in range(sheet.min_row+1,sheet.max_row+1):
			sheet.row_dimensions[x].height = LONG_ROW
	# wrap text for long columns
	for column_cells in sheet.columns:
		# look for columns larger than LONG_COLUMN to wrap it out
		if sheet.column_dimensions[get_column_letter(column_cells[0].column)].width >= LONG_COLUMN:
			col=column_cells[0].column
			# set wrap text for each colum
			for rows in sheet.iter_rows(min_row=sheet.min_row+1, max_row=sheet.max_row, min_col=col, max_col=col):
				for cell in rows:
					cell.alignment = Alignment(wrapText=True)
	workbook.save(filename=filename)

def export_csv(data, filename):
	'''Given a dataset (as a list of lists), save it to the specified file
	as a csv file'''
	with open(filename, "w", newline="") as f:
		writer = csv.writer(f)
		writer.writerows(data)

def export_txt(data):
	'''Given a dataset (as a list of lists), print it to the stdout'''
	for row in data:
		print(row)

def list_of_components(conf=None):
	'''Print the list of components of the given configuration file'''
	data={}
	# if not set, use the default component configuration file
	conf=os.path.join(get_exec_path(),'config','components.yaml') if conf is None else conf
	# load component configuration
	with open(conf) as f:
		data = yaml.load(f, Loader=yaml.FullLoader)
	print('''The list of components defined to extract from exported Infa XML files with its 
associated command line options are shown below.
Components:''')
	for k,v in data['components'].items():
		print(f'  {k}\t--extract {k}')
	print('''
The following components are nested by hierachy level and its associated command line 
options are shown below.
Nested components:''')
	for k,v in data['nested'].items():
		print(f'  {k}:')
		for kitm, vitm in v.items():
			print(f'    {kitm}\t--from {k} --extract {kitm}')
